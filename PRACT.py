import time

from bs4 import BeautifulSoup
import requests
import openpyxl

# URL страницы поиска
url_ctl = 'https://www.chipdip.ru/search?searchtext='

# Файл товаров
file = "Reestr.xlsx"

# Инициализируем библиотеку для работы с Excel и считываем значения колонок
wb_obj = openpyxl.load_workbook(file)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

for i in range(1, m_row + 1):
    f = open("output.txt", "a")
    # Считываем значение ячейки
    cell_obj = sheet_obj.cell(row=i + 1, column=1)
    print('-------------------------', file=f)
    print('Товар: ' + str(cell_obj.value), file=f)
    # Формируем URL с поиском переменная url_src + значение ячейки
    url = (url_ctl + str(cell_obj.value))
    # Делаем http запрос библиотекой requests и помещаем ответ сервера в переменную r
    r = requests.get(url, headers={
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/103.0.5060.53 Safari/537.36 Edg/103.0.1264.37'})
    # Инициализируем парсер BeautifulSoup
    soup = BeautifulSoup(r.text, 'html.parser')
    # Находим на полученной странице с сайта элемент с классом content(именно в в нём содержится информация о товаре)
    url_item = soup.find_all('div', class_="content")
    # Ищем нужную нам информацию в подклассах и записываем её в контейнер
    products = []
    try:
        for item in url_item:
            products.append({
                'Title': item.find('td', class_="h_name").get_text(strip=True),
                'Price': item.find('span', class_="price").get_text(strip=True).translate(
                    {ord(i): None for i in '\xa0'})
            })
        time.sleep(1.5)
        print(products, file=f)
    except:
        time.sleep(1.5)
        print("Товара нет в наличии", file=f)
    f.close()
    print("writing to a file....Please,wait")
